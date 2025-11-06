from openpyxl import Workbook, load_workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "Original"
ws1.append(["A", "B", "C"])

ws2 = wb.create_sheet("Copy")
for row in ws1.iter_rows(values_only=True):
    ws2.append(row)

wb.save("copy.xlsx")
