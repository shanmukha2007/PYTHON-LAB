from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.append(["Product", "Qty", "Price", "Total"])
ws.append(["Laptop", 2, 50000, None])
ws.append(["Mouse", 5, 500, None])

for row in ws.iter_rows(min_row=2, max_col=4):
    qty, price = row[1].value, row[2].value
    ws.cell(row=row[0].row, column=4, value=qty * price)
wb.save("sales.xlsx")