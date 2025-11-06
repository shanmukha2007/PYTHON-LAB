from openpyxl import Workbook, load_workbook

# 1. Read marks &amp; write average
wb = Workbook()
ws = wb.active
ws.append(["Name", "Mark1", "Mark2", "Mark3"])
ws.append(["Ravi", 80, 90, 85])
ws.append(["Anita", 70, 75, 80])

for row in ws.iter_rows(min_row=2, max_col=4):
    marks = [cell.value for cell in row[1:]]
    avg = sum(marks) / len(marks)
    ws.cell(row=row[0].row, column=5, value=avg)

wb.save("marks.xlsx")